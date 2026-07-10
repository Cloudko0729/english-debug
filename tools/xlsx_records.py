# 擷取 xlsx 的 Engineering Record 欄（J）→ 純文字（私有，存 P 槽）
# 用法: python tools/xlsx_records.py "P:/AI/claude/Project View __ Non-ATE_2026.xlsx" "P:/AI/claude/eng_records.txt"
import sys
from pathlib import Path

import openpyxl

wb = openpyxl.load_workbook(sys.argv[1], read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    v = row[9]  # J 欄 Engineering Record
    if v is None:
        continue
    s = str(v).strip()
    if s:
        records.append(s)

Path(sys.argv[2]).write_text("\n=====\n".join(records), encoding="utf-8")
total = sum(len(r) for r in records)
print(f"records: {len(records)} 筆, 共 {total//1024}KB")
uniq = len(set(records))
print(f"去重後: {uniq} 筆")
